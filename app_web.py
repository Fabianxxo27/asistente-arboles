import streamlit as st
import gspread
from google.oauth2.service_account import Credentials
from datetime import datetime
import json
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from io import BytesIO
import copy
import tempfile
import shutil

# Configuración de la página
st.set_page_config(
    page_title="🌳 Asistente de Registro - Árboles",
    page_icon="🌳",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# CSS personalizado
st.markdown("""
<style>
    .stApp {
        max-width: 1200px;
        margin: 0 auto;
    }
    .header-style {
        background: linear-gradient(135deg, #2c5f2d 0%, #3d7f3f 100%);
        padding: 20px;
        border-radius: 10px;
        color: white;
        text-align: center;
        margin-bottom: 30px;
    }
    .section-header {
        background-color: #2c5f2d;
        color: white;
        padding: 10px;
        border-radius: 5px;
        margin-top: 20px;
        margin-bottom: 10px;
        font-weight: bold;
    }
    div[data-testid="stCheckbox"] {
        display: inline-block;
        margin-right: 10px;
    }
    .success-box {
        background-color: #d4edda;
        border: 1px solid #c3e6cb;
        color: #155724;
        padding: 15px;
        border-radius: 5px;
        margin: 10px 0;
    }
</style>
""", unsafe_allow_html=True)

# Función para conectar con Google Sheets
@st.cache_resource
def conectar_google_sheets():
    """Conecta a Google Sheets usando credenciales"""
    try:
        # Intenta cargar credenciales desde archivo
        if os.path.exists('credenciales.json'):
            SCOPES = [
                'https://www.googleapis.com/auth/spreadsheets',
                'https://www.googleapis.com/auth/drive'
            ]
            creds = Credentials.from_service_account_file('credenciales.json', scopes=SCOPES)
            client = gspread.authorize(creds)
            return client
        # Si no existe archivo, intenta cargar desde secrets de Streamlit
        elif 'gcp_service_account' in st.secrets:
            SCOPES = [
                'https://www.googleapis.com/auth/spreadsheets',
                'https://www.googleapis.com/auth/drive'
            ]
            creds = Credentials.from_service_account_info(
                st.secrets["gcp_service_account"],
                scopes=SCOPES
            )
            client = gspread.authorize(creds)
            return client
        else:
            return None
    except Exception as e:
        st.error(f"Error al conectar: {str(e)}")
        return None

def obtener_ultimo_codigo(worksheet):
    """Obtiene el último código de la columna C"""
    try:
        valores = worksheet.col_values(3)  # Columna C
        # Filtrar solo números
        codigos = [int(v) for v in valores[1:] if v and str(v).isdigit()]
        if codigos:
            return max(codigos)
        return 19222  # Valor inicial si no hay códigos
    except:
        return 19222

def agregar_fila_excel(worksheet_excel, datos):
    """Agrega o actualiza una fila en Excel según el código/ID"""
    try:
        codigo_a_escribir = str(datos.get('codigo', ''))
        
        # Buscar si el código ya existe
        fila_objetivo = None
        for idx, row in enumerate(worksheet_excel.iter_rows(min_row=2, min_col=3, max_col=3), start=2):
            if str(row[0].value).strip() == codigo_a_escribir.strip():
                fila_objetivo = idx
                break
        
        # Si no existe, agregar al final
        if fila_objetivo is None:
            fila_objetivo = worksheet_excel.max_row + 1
        
        # Escribir datos básicos
        worksheet_excel.cell(fila_objetivo, 1).value = datos.get('entidad', 'OTRO')  # Columna A
        worksheet_excel.cell(fila_objetivo, 2).value = datos.get('nit', '901145808-5')  # Columna B
        worksheet_excel.cell(fila_objetivo, 3).value = codigo_a_escribir  # Columna C
        
        # Estado Físico Fuste
        for col_excel, marcado in datos.get('checks_fuste', {}).items():
            if marcado:
                worksheet_excel.cell(fila_objetivo, col_excel).value = '1'
        
        # Estados generales
        if datos.get('fuste_general'):
            worksheet_excel.cell(fila_objetivo, 23).value = datos['fuste_general']
        if datos.get('raiz_especifico'):
            worksheet_excel.cell(fila_objetivo, 24).value = datos['raiz_especifico']
        if datos.get('raiz_general'):
            worksheet_excel.cell(fila_objetivo, 25).value = datos['raiz_general']
        
        # Estado Sanitario Copa
        for col_excel, marcado in datos.get('checks_copa', {}).items():
            if marcado:
                worksheet_excel.cell(fila_objetivo, col_excel).value = '1'
        
        # Estado Sanitario Fuste
        for col_excel, marcado in datos.get('checks_fuste_san', {}).items():
            if marcado:
                worksheet_excel.cell(fila_objetivo, col_excel).value = '1'
        
        # Estado Sanitario Raíz Específico
        if datos.get('san_raiz_especifico'):
            worksheet_excel.cell(fila_objetivo, 48).value = datos['san_raiz_especifico']
        
        # Estados sanitarios generales
        if datos.get('san_general'):
            worksheet_excel.cell(fila_objetivo, 49).value = datos['san_general']
        if datos.get('san_copa_general'):
            worksheet_excel.cell(fila_objetivo, 50).value = datos['san_copa_general']
        if datos.get('san_fuste_general'):
            worksheet_excel.cell(fila_objetivo, 51).value = datos['san_fuste_general']
        if datos.get('san_raiz_general'):
            worksheet_excel.cell(fila_objetivo, 52).value = datos['san_raiz_general']
        
        # Interferencia con líneas de servicios
        for col_excel, marcado in datos.get('checks_servicios', {}).items():
            if marcado:
                worksheet_excel.cell(fila_objetivo, col_excel).value = '1'
        
        # Causas de Poda
        for col_excel, marcado in datos.get('checks_poda', {}).items():
            if marcado:
                worksheet_excel.cell(fila_objetivo, col_excel).value = '1'
        
        # Tipo e intensidad poda
        if datos.get('tipo_poda'):
            worksheet_excel.cell(fila_objetivo, 66).value = datos['tipo_poda']
        if datos.get('intensidad'):
            worksheet_excel.cell(fila_objetivo, 67).value = datos['intensidad']
        
        # Concepto Técnico
        for col_excel, marcado in datos.get('checks_concepto', {}).items():
            if marcado:
                worksheet_excel.cell(fila_objetivo, col_excel).value = '1'
        
        # Residuos
        if datos.get('residuos'):
            worksheet_excel.cell(fila_objetivo, 77).value = datos['residuos']
        
        return True, fila_objetivo
    except Exception as e:
        return False, str(e)

def agregar_fila_sheets(worksheet, datos):
    """Agrega o actualiza una fila en Google Sheets según el código/ID"""
    try:
        # Obtener todos los valores de la columna C (códigos)
        valores_columna_c = worksheet.col_values(3)  # Columna C
        codigo_a_escribir = str(datos.get('codigo', ''))
        
        # Buscar si el código ya existe
        fila_objetivo = None
        for idx, valor in enumerate(valores_columna_c, start=1):
            if str(valor).strip() == codigo_a_escribir.strip():
                fila_objetivo = idx
                break
        
        # Si no existe, agregar al final
        if fila_objetivo is None:
            valores = worksheet.get_all_values()
            fila_objetivo = len(valores) + 1
        
        # Preparar fila con 80 columnas (ajusta según tu Excel)
        fila_datos = [''] * 80
        
        # Datos básicos
        fila_datos[0] = datos.get('entidad', 'OTRO')  # Columna A
        fila_datos[1] = datos.get('nit', '901145808-5')  # Columna B
        fila_datos[2] = datos.get('codigo', '')  # Columna C
        
        # Estado Físico Fuste (columnas 4-19, índices 3-18)
        for col_excel, marcado in datos.get('checks_fuste', {}).items():
            if marcado:
                fila_datos[col_excel - 1] = '1'
        
        # Estados generales
        if datos.get('fuste_general'):
            fila_datos[22] = datos['fuste_general']  # Columna W (23)
        if datos.get('raiz_especifico'):
            fila_datos[23] = datos['raiz_especifico']  # Columna X (24)
        if datos.get('raiz_general'):
            fila_datos[24] = datos['raiz_general']  # Columna Y (25)
        
        # Estado Sanitario Copa (columnas 26-40)
        for col_excel, marcado in datos.get('checks_copa', {}).items():
            if marcado:
                fila_datos[col_excel - 1] = '1'
        
        # Estado Sanitario Fuste (columnas 41-47)
        for col_excel, marcado in datos.get('checks_fuste_san', {}).items():
            if marcado:
                fila_datos[col_excel - 1] = '1'
        
        # Estado Sanitario Raíz Específico
        if datos.get('san_raiz_especifico'):
            fila_datos[47] = datos['san_raiz_especifico']  # Columna 48
        
        # Estados sanitarios generales
        if datos.get('san_general'):
            fila_datos[48] = datos['san_general']  # Columna 49
        if datos.get('san_copa_general'):
            fila_datos[49] = datos['san_copa_general']  # Columna 50
        if datos.get('san_fuste_general'):
            fila_datos[50] = datos['san_fuste_general']  # Columna 51
        if datos.get('san_raiz_general'):
            fila_datos[51] = datos['san_raiz_general']  # Columna 52
        
        # Interferencia con líneas de servicios (columnas 53-56)
        for col_excel, marcado in datos.get('checks_servicios', {}).items():
            if marcado:
                fila_datos[col_excel - 1] = '1'
        
        # Causas de Poda (columnas 57-65)
        for col_excel, marcado in datos.get('checks_poda', {}).items():
            if marcado:
                fila_datos[col_excel - 1] = '1'
        
        # Tipo e intensidad poda
        if datos.get('tipo_poda'):
            fila_datos[65] = datos['tipo_poda']  # Columna 66
        if datos.get('intensidad'):
            fila_datos[66] = datos['intensidad']  # Columna 67
        
        # Concepto Técnico (columnas 68-76)
        for col_excel, marcado in datos.get('checks_concepto', {}).items():
            if marcado:
                fila_datos[col_excel - 1] = '1'
        
        # Residuos
        if datos.get('residuos'):
            fila_datos[76] = datos['residuos']  # Columna 77
        
        # Escribir en la fila específica (reemplaza si existe, crea si no)
        # Construir el rango A:fila_objetivo hasta la columna que necesites
        rango = f'A{fila_objetivo}:BX{fila_objetivo}'  # BX = columna 76, ajusta si necesitas más
        worksheet.update(rango, [fila_datos[:76]])  # Escribir solo las columnas necesarias
        
        return True, fila_objetivo
    except Exception as e:
        return False, str(e)

# Header
st.markdown("""
<div class="header-style">
    <h1>🌳 Asistente de Registro de Árboles</h1>
    <p>Registro en tiempo real sincronizado con Google Sheets o Excel</p>
</div>
""", unsafe_allow_html=True)

# Selector de modo
st.markdown("### 📂 Selecciona el modo de trabajo:")
modo = st.radio(
    "¿Cómo quieres trabajar?",
    ["🌐 Google Sheets (en la nube)", "📁 Archivo Excel (subir/descargar)"],
    horizontal=True
)

# Variable global para saber qué modo está activo
usa_google_sheets = (modo == "🌐 Google Sheets (en la nube)")

# ============= MODO GOOGLE SHEETS =============
if usa_google_sheets:
    # Intentar conectar
    client = conectar_google_sheets()

    if client is None:
        st.error("⚠️ **No se pudo conectar a Google Sheets**")
        st.info("""
        **Configuración necesaria:**
        
        1. Crea un proyecto en Google Cloud Platform
        2. Habilita Google Sheets API y Google Drive API
        3. Crea credenciales de cuenta de servicio
        4. Descarga el archivo JSON y guárdalo como `credenciales.json` en esta carpeta
        5. Comparte tu Google Sheet con el email de la cuenta de servicio
        
        📖 Ver instrucciones completas en `README_WEB.md`
        """)
        st.stop()

    # Selector de hoja
    spreadsheet_id = st.text_input(
        "🔗 ID de Google Sheets",
        placeholder="Pega aquí el ID de tu Google Sheet (desde la URL)",
        help="El ID está en la URL: https://docs.google.com/spreadsheets/d/[ID]/edit"
    )

    if not spreadsheet_id:
        st.warning("⬆️ Ingresa el ID de tu Google Sheet para continuar")
        st.stop()

    try:
        spreadsheet = client.open_by_key(spreadsheet_id)
        
        # Buscar la hoja BASE DE DATOS (con o sin espacios)
        worksheet = None
        for sheet in spreadsheet.worksheets():
            if "BASE DE DATOS" in sheet.title.upper():
                worksheet = sheet
                break
        
        if worksheet is None:
            st.error("❌ No se encontró ninguna hoja con nombre 'BASE DE DATOS'")
            st.info(f"Hojas disponibles: {', '.join([s.title for s in spreadsheet.worksheets()])}")
            st.stop()
        
        st.success(f"✅ **Conectado a:** {spreadsheet.title} (Hoja: {worksheet.title})")
        
        # Obtener último código
        ultimo_codigo = obtener_ultimo_codigo(worksheet)
        siguiente_codigo = ultimo_codigo + 1
        
    except Exception as e:
        st.error(f"❌ Error: {str(e)}")
        st.info("Verifica que:\n- El ID sea correcto\n- Hayas compartido el sheet con: asistente@excel-485919.iam.gserviceaccount.com\n- La cuenta de servicio tenga permisos de Editor")
        st.stop()

    # Inicializar session state para el código
    if 'codigo_actual' not in st.session_state:
        st.session_state.codigo_actual = siguiente_codigo

    # Variable para controlar si se debe reiniciar el formulario
    if 'form_key' not in st.session_state:
        st.session_state.form_key = 0
        
# ============= MODO ARCHIVO EXCEL =============
else:
    st.markdown("### 📤 Sube tu archivo Excel")
    
    uploaded_file = st.file_uploader(
        "Selecciona tu archivo Excel (.xlsx o .xlsm)",
        type=['xlsx', 'xlsm'],
        help="Sube el archivo Excel que contiene la hoja 'BASE DE DATOS'"
    )
    
    if uploaded_file is None:
        st.info("⬆️ Sube tu archivo Excel para continuar")
        st.stop()
    
    # Cargar Excel en memoria
    try:
        if 'excel_workbook' not in st.session_state or st.session_state.get('uploaded_filename') != uploaded_file.name:
            # Guardar el archivo original completo en bytes
            uploaded_file.seek(0)
            st.session_state.excel_original_bytes = uploaded_file.read()
            
            # Cargar workbook desde bytes (no desde uploaded_file que puede cerrarse)
            temp_file = BytesIO(st.session_state.excel_original_bytes)
            
            try:
                # Cargar manteniendo VBA, fórmulas, todo
                wb = load_workbook(
                    temp_file, 
                    keep_vba=True if uploaded_file.name.endswith('.xlsm') else False,
                    data_only=False,
                    keep_links=True
                )
            except Exception:
                # Si falla, intentar sin VBA
                temp_file = BytesIO(st.session_state.excel_original_bytes)
                wb = load_workbook(
                    temp_file,
                    data_only=False,
                    keep_links=True
                )
            
            # NO modificar ni eliminar NADA del workbook
            st.session_state.excel_workbook = wb
            st.session_state.uploaded_filename = uploaded_file.name
            st.session_state.registros_agregados = 0
            st.session_state.datos_agregados = []
        else:
            wb = st.session_state.excel_workbook
        
        # Buscar hoja BASE DE DATOS
        worksheet_excel = None
        for sheet_name in wb.sheetnames:
            if "BASE DE DATOS" in sheet_name.upper():
                worksheet_excel = wb[sheet_name]
                break
        
        if worksheet_excel is None:
            st.error("❌ No se encontró ninguna hoja con nombre 'BASE DE DATOS'")
            st.info(f"Hojas disponibles: {', '.join(wb.sheetnames)}")
            st.stop()
        
        # Obtener último código del Excel
        ultimo_codigo_excel = 19222
        for row in worksheet_excel.iter_rows(min_row=2, min_col=3, max_col=3):
            valor = row[0].value
            if valor and str(valor).isdigit():
                ultimo_codigo_excel = max(ultimo_codigo_excel, int(valor))
        
        siguiente_codigo = ultimo_codigo_excel + 1
        
        st.success(f"✅ **Archivo cargado:** {uploaded_file.name} (Hoja: {worksheet_excel.title})")
        
        # Mostrar estadísticas
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("📊 Total filas", worksheet_excel.max_row - 1)
        with col2:
            st.metric("🆕 Registros agregados", st.session_state.registros_agregados)
        with col3:
            st.metric("🔢 Último código", ultimo_codigo_excel)
        
        if 'codigo_actual' not in st.session_state:
            st.session_state.codigo_actual = siguiente_codigo
        
        # Variable para controlar si se debe reiniciar el formulario
        if 'form_key' not in st.session_state:
            st.session_state.form_key = 0
        
        worksheet = worksheet_excel  # Para usar en el formulario
        
    except Exception as e:
        st.error(f"❌ Error al cargar Excel: {str(e)}")
        st.stop()

# Formulario
with st.form(key=f"formulario_arbol_{st.session_state.form_key}"):
    
    # Datos básicos
    col1, col2, col3 = st.columns(3)
    with col1:
        entidad = st.text_input("Entidad:", value="OTRO")
    with col2:
        nit = st.text_input("NIT:", value="901145808-5")
    with col3:
        codigo = st.number_input("ID/Código:", value=st.session_state.codigo_actual, step=1)
    
    # Estado Físico Fuste
    st.markdown('<div class="section-header">🌳 Estado Físico Fuste</div>', unsafe_allow_html=True)
    cols_fuste = st.columns(6)
    checks_fuste = {}
    opciones_fuste = [
        ("B", 4), ("Bb", 5), ("BB", 6), ("FR", 7), ("I", 8), ("MI", 9),
        ("To", 10), ("C", 11), ("Rv", 12), ("Ac", 13), ("An", 14), ("Dc", 15),
        ("SB", 16), ("Ag", 17), ("Poe", 18), ("Pe", 19), ("DM-L", 20), ("DM-M", 21), ("DM-G", 22)
    ]
    for idx, (nombre, col) in enumerate(opciones_fuste):
        with cols_fuste[idx % 6]:
            checks_fuste[col] = st.checkbox(nombre, key=f"fuste_{col}")
    
    col1, col2, col3 = st.columns(3)
    with col1:
        fuste_general = st.selectbox("Estado Fuste General:", 
            ["", "Bueno", "Regular", "Malo", "Suprimido"], index=1)
    with col2:
        raiz_especifico = st.selectbox("Estado Raíz Específico:", 
            ["", "No apreciable", "Visible", "Superficial", "Profunda"], index=1)
    with col3:
        raiz_general = st.selectbox("Estado Raíz General:", 
            ["", "Bueno", "Regular", "Malo"], index=1)
    
    # Estado Sanitario Copa (15 opciones: columnas 26-40)
    st.markdown('<div class="section-header">🍃 Estado Sanitario Específico Copa</div>', unsafe_allow_html=True)
    cols_copa = st.columns(6)
    checks_copa = {}
    opciones_copa = [
        ("He", 26), ("An", 27), ("Ag", 28), ("Ne", 29), ("Tu", 30),
        ("Cl", 31), ("Ma", 32), ("Ca", 33), ("PL", 34), ("Mi", 35),
        ("C", 36), ("Ro", 37), ("Psu", 38), ("PI", 39), ("NA", 40)
    ]
    for idx, (nombre, col) in enumerate(opciones_copa):
        with cols_copa[idx % 6]:
            checks_copa[col] = st.checkbox(nombre, key=f"copa_{col}")
    
    # Estado Sanitario Fuste (7 opciones: columnas 41-47)
    st.markdown('<div class="section-header">🏥 Estado Sanitario Específico Fuste</div>', unsafe_allow_html=True)
    cols_fuste_san = st.columns(7)
    checks_fuste_san = {}
    opciones_fuste_san = [
        ("Ch", 41), ("Plf", 42), ("Go", 43), ("Tu", 44),
        ("Ag", 45), ("PI", 46), ("NA", 47)
    ]
    for idx, (nombre, col) in enumerate(opciones_fuste_san):
        with cols_fuste_san[idx]:
            checks_fuste_san[col] = st.checkbox(nombre, key=f"fuste_san_{col}")
    
    # Estado Sanitario Raíz Específico (columna 48)
    san_raiz_especifico = st.text_input("Estado Sanitario Raíz Específico:", value="Ninguna de las anteriores")
    
    # Estados sanitarios generales (columnas 49-52)
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        san_general = st.selectbox("Estado Sanitario General:", 
            ["", "Bueno", "Regular", "Malo"], index=1)
    with col2:
        san_copa_general = st.selectbox("Estado Sanitario Copa General:", 
            ["", "Bueno", "Regular", "Malo"], index=1)
    with col3:
        san_fuste_general = st.selectbox("Estado Sanitario Fuste General:", 
            ["", "Bueno", "Regular", "Malo"], index=1)
    with col4:
        san_raiz_general = st.selectbox("Estado Sanitario Raíz General:", 
            ["", "Bueno", "Regular", "Malo"], index=1)
    
    # Interferencia con líneas de servicios (columnas 53-56)
    st.markdown('<div class="section-header">⚡ Interferencia con Líneas de Servicios</div>', unsafe_allow_html=True)
    cols_servicios = st.columns(4)
    checks_servicios = {}
    opciones_servicios = [
        ("Luminarias", 53), ("Alta tensión", 54), ("Media Tensión", 55), ("Subterráneas", 56)
    ]
    for idx, (nombre, col) in enumerate(opciones_servicios):
        with cols_servicios[idx]:
            checks_servicios[col] = st.checkbox(nombre, key=f"servicio_{col}")
    
    # Causas de intervención de la poda (columnas 57-65)
    st.markdown('<div class="section-header">✂️ Causas de Intervención de la Poda</div>', unsafe_allow_html=True)
    cols_poda = st.columns(5)
    checks_poda = {}
    opciones_poda = [
        ("Corteza incluida", 57), ("Grietas", 58), ("Excesivos rebrotes", 59),
        ("Pudriciones", 60), ("Ramas rotas o muertas", 61), ("Copa asimétrica", 62),
        ("Ramas sobreextendidas", 63), ("Ramas pendulares", 64), ("Raíces extranguladoras", 65)
    ]
    for idx, (nombre, col) in enumerate(opciones_poda):
        with cols_poda[idx % 5]:
            checks_poda[col] = st.checkbox(nombre, key=f"poda_{col}")
    
    col1, col2, col3 = st.columns(3)
    with col1:
        tipo_poda = st.selectbox("Tipo de Poda:", 
            ["", "De mejoramiento-Estructura", "De mantenimiento", "Especial", "Sanitaria"], 
            index=1)
    with col2:
        intensidad = st.text_input("Intensidad de la Poda (%):")
    with col3:
        residuos = st.text_input("Residuos Generados (kg):")
    
    # Concepto Técnico (columnas 68-76: 9 opciones)
    st.markdown('<div class="section-header">📋 Concepto Técnico</div>', unsafe_allow_html=True)
    cols_concepto = st.columns(5)
    checks_concepto = {}
    opciones_concepto = [
        ("Inclinación severa hacia estructuras", 68),
        ("Problemas de seguridad en la zona", 69),
        ("Minimizar riesgo de volcamiento", 70),
        ("Mantenimiento de individuos jóvenes", 71),
        ("Disminuir competencia con otros individuos arbóreos", 72),
        ("Mantenimiento de arbolado adulto", 73),
        ("Despeje de cono luminico", 74),
        ("Liberación de infraestructura urbana y movilidad", 75),
        ("Despeje sistema circulación urbana", 76)
    ]
    for idx, (nombre, col) in enumerate(opciones_concepto):
        with cols_concepto[idx % 5]:
            checks_concepto[col] = st.checkbox(nombre, key=f"concepto_{col}")
    
    # Botón enviar
    st.markdown("<br>", unsafe_allow_html=True)
    texto_boton = "➕ Agregar Registro a Google Sheets" if usa_google_sheets else "➕ Agregar Registro al Excel"
    submitted = st.form_submit_button(texto_boton, use_container_width=True, type="primary")

# Procesar formulario FUERA del with st.form
if submitted:
    # Preparar datos
    datos = {
        'entidad': entidad,
        'nit': nit,
        'codigo': str(codigo),
        'checks_fuste': {col: True for col, val in checks_fuste.items() if val},
        'fuste_general': fuste_general,
        'raiz_especifico': raiz_especifico,
        'raiz_general': raiz_general,
        'checks_copa': {col: True for col, val in checks_copa.items() if val},
        'checks_fuste_san': {col: True for col, val in checks_fuste_san.items() if val},
        'san_raiz_especifico': san_raiz_especifico,
        'san_general': san_general,
        'san_copa_general': san_copa_general,
        'san_fuste_general': san_fuste_general,
        'san_raiz_general': san_raiz_general,
        'checks_servicios': {col: True for col, val in checks_servicios.items() if val},
        'checks_poda': {col: True for col, val in checks_poda.items() if val},
        'tipo_poda': tipo_poda,
        'intensidad': intensidad,
        'residuos': residuos,
        'checks_concepto': {col: True for col, val in checks_concepto.items() if val}
    }
    
    # Guardar según el modo
    if usa_google_sheets:
        with st.spinner('Guardando en Google Sheets...'):
            exito, resultado = agregar_fila_sheets(worksheet, datos)
    else:
        with st.spinner('Guardando en Excel...'):
            exito, resultado = agregar_fila_excel(worksheet, datos)
            if exito:
                # Incrementar contador de registros agregados
                st.session_state.registros_agregados += 1
                # Guardar datos para CSV alternativo
                if 'datos_agregados' not in st.session_state:
                    st.session_state.datos_agregados = []
                st.session_state.datos_agregados.append({
                    'fila': resultado,
                    'id': codigo,
                    'datos': datos
                })
    
    if exito:
        # Auto-incrementar código para el siguiente
        st.session_state.codigo_actual = int(codigo) + 1
        # Cambiar la key del formulario para forzar reset
        st.session_state.form_key += 1
        
        st.success(f"✅ **Registro guardado en fila {resultado}** (ID: {codigo})")
        
        # Mensaje especial para modo Excel
        if not usa_google_sheets:
            st.info(f"💾 **{st.session_state.registros_agregados} registro(s) en memoria.** Usa el botón 'Descargar Excel' cuando termines.")
        
        st.balloons()
        st.rerun()
    else:
        st.error(f"❌ Error al guardar: {resultado}")

# Sección de descarga FUERA del formulario (solo para modo Excel)
if not usa_google_sheets:
    # Guardar workbook en bytes
    st.markdown("---")
    
    col_desc, col_btn = st.columns([3, 2])
    
    with col_desc:
        if st.session_state.registros_agregados > 0:
            st.success(f"✅ **{st.session_state.registros_agregados} registro(s) agregado(s)**")
            st.caption("Descarga el archivo completo con todos tus cambios")
        else:
            st.info("ℹ️ Aún no has agregado registros")
    
    with col_btn:
        # Determinar extensión según el archivo original
        extension = "xlsm" if st.session_state.uploaded_filename.endswith('.xlsm') else "xlsx"
        
        # ESTRATEGIA: Reconstruir desde bytes originales + aplicar cambios
        try:
            # Cargar copia fresca del archivo original cada vez
            temp_file = BytesIO(st.session_state.excel_original_bytes)
            
            # Cargar con keep_vba
            wb_download = load_workbook(
                temp_file,
                keep_vba=True if extension == "xlsm" else False,
                data_only=False,
                keep_links=True
            )
            
            # Encontrar la hoja BASE DE DATOS
            worksheet_download = None
            for sheet_name in wb_download.sheetnames:
                if "BASE DE DATOS" in sheet_name.upper():
                    worksheet_download = wb_download[sheet_name]
                    break
            
            if worksheet_download:
                # Obtener la hoja del workbook en sesión que tiene los datos actualizados
                worksheet_session = None
                for sheet_name in st.session_state.excel_workbook.sheetnames:
                    if "BASE DE DATOS" in sheet_name.upper():
                        worksheet_session = st.session_state.excel_workbook[sheet_name]
                        break
                
                if worksheet_session:
                    # Copiar solo los VALORES de las celdas modificadas
                    # Obtener el rango de filas que tienen datos
                    max_row = worksheet_session.max_row
                    max_col = worksheet_session.max_column
                    
                    # Copiar valores desde la fila 2 (después de encabezados)
                    for row_idx in range(2, max_row + 1):
                        for col_idx in range(1, max_col + 1):
                            # Verificar que no sea una celda combinada antes de escribir
                            cell_download = worksheet_download.cell(row_idx, col_idx)
                            
                            # Solo escribir si NO es MergedCell
                            if not isinstance(cell_download, MergedCell):
                                valor = worksheet_session.cell(row_idx, col_idx).value
                                cell_download.value = valor
            
            # Guardar el workbook reconstruido
            output = BytesIO()
            wb_download.save(output)
            wb_download.close()
            output.seek(0)
            excel_data = output.getvalue()
            
            # Verificar tamaño
            if len(excel_data) > 1000:
                st.download_button(
                    label=f"📥 Descargar Excel Completo",
                    data=excel_data,
                    file_name=f"{st.session_state.uploaded_filename.replace('.xlsx', '').replace('.xlsm', '')}_actualizado_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{extension}",
                    mime="application/vnd.ms-excel.sheet.macroEnabled.12" if extension == "xlsm" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary"
                )
                st.caption("💾 Archivo actualizado con todos los nuevos registros")
                st.success(f"✅ Listo para descargar con {st.session_state.registros_agregados} registro(s) nuevo(s)")
            else:
                raise Exception("Archivo muy pequeño, posible error")
                
        except Exception as e:
            # Si falla, mostrar el error y ofrecer alternativas
            st.error(f"❌ Error al guardar: {str(e)}")
            st.warning("⚠️ El archivo tiene elementos que impiden el guardado automático")
            
            # Ofrecer alternativas
            excel_guardado = False
    
    # Si NO se pudo guardar el archivo original, ofrecer alternativas
    if 'excel_guardado' in locals() and excel_guardado == False:
        st.markdown("### 📄 Opciones de Descarga Alternativas")
        st.info("💡 Tu archivo original tiene elementos complejos. Elige una opción:")
        
        with st.expander("✅ RECOMENDADO: Descargar Excel nuevo con TODOS los datos", expanded=True):
            try:
                from openpyxl import Workbook
                
                # Crear nuevo workbook
                nuevo_wb = Workbook()
                nuevo_ws = nuevo_wb.active
                nuevo_ws.title = "BASE DE DATOS"
                
                # Copiar solo los datos (valores) de la hoja original
                worksheet_excel = None
                for sheet_name in st.session_state.excel_workbook.sheetnames:
                    if "BASE DE DATOS" in sheet_name.upper():
                        worksheet_excel = st.session_state.excel_workbook[sheet_name]
                        break
                
                if worksheet_excel:
                    # Copiar datos celda por celda (solo valores, sin formato)
                    max_row = worksheet_excel.max_row
                    max_col = worksheet_excel.max_column
                    
                    st.info(f"📊 Copiando {max_row} filas y {max_col} columnas...")
                    
                    for row_idx, row in enumerate(worksheet_excel.iter_rows(min_row=1, max_row=max_row, max_col=max_col), 1):
                        for col_idx, cell in enumerate(row, 1):
                            nuevo_ws.cell(row=row_idx, column=col_idx, value=cell.value)
                    
                    # Guardar el nuevo workbook
                    nuevo_output = BytesIO()
                    nuevo_wb.save(nuevo_output)
                    nuevo_output.seek(0)
                    nuevo_data = nuevo_output.getvalue()
                    
                    st.success("✅ Excel nuevo creado exitosamente")
                    st.download_button(
                        label="📥 Descargar Excel Nuevo (TODOS los datos)",
                        data=nuevo_data,
                        file_name=f"arboles_completo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        type="primary"
                    )
                    st.caption(f"✅ Contiene TODOS los datos ({max_row} filas) sin imágenes/formatos complejos")
                else:
                    st.error("No se encontró la hoja BASE DE DATOS")
            except Exception as e2:
                st.error(f"❌ Error al crear Excel nuevo: {str(e2)}")
        
        with st.expander("📋 Opción 2: Descargar solo nuevos registros en CSV"):
                if st.session_state.registros_agregados > 0 and 'datos_agregados' in st.session_state:
                    # Crear CSV con los datos agregados
                    import csv
                    csv_output = BytesIO()
                    csv_output.write('\ufeff'.encode('utf-8'))  # BOM para Excel
                    
                    # Escribir encabezados
                    headers = ['Fila', 'Entidad', 'NIT', 'ID/Código', 'Estado Fuste General', 'Estado Raíz General', 
                              'Estado Sanitario General', 'Tipo Poda', 'Intensidad', 'Residuos']
                    
                    csv_text = ','.join(headers) + '\n'
                    
                    for item in st.session_state.datos_agregados:
                        d = item['datos']
                        row = [
                            str(item['fila']),
                            d.get('entidad', ''),
                            d.get('nit', ''),
                            d.get('codigo', ''),
                            d.get('fuste_general', ''),
                            d.get('raiz_general', ''),
                            d.get('san_general', ''),
                            d.get('tipo_poda', ''),
                            d.get('intensidad', ''),
                            d.get('residuos', '')
                        ]
                        csv_text += ','.join([f'"{r}"' for r in row]) + '\n'
                    
                    csv_output.write(csv_text.encode('utf-8'))
                    csv_data = csv_output.getvalue()
                    
                    st.download_button(
                        label="📥 Descargar Registros Nuevos (CSV)",
                        data=csv_data,
                        file_name=f"registros_nuevos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                        mime="text/csv",
                        use_container_width=True
                    )
                    
                    st.caption(f"📋 {st.session_state.registros_agregados} registro(s) para copiar a tu Excel original")
                else:
                    st.info("No hay registros nuevos para exportar")
        
        # Mostrar recomendación
        st.markdown("---")
        st.info("""
        **💡 Recomendación final:**
        - **Opción 1 (Excel nuevo)**: Archivo completo listo para usar ✅
        - **Usa Google Sheets**: Nunca tendrás estos problemas 🌐
        """)
        
        if st.checkbox("🔍 Ver detalles técnicos"):
            st.code(f"Archivo: {st.session_state.uploaded_filename}\nError: Openpyxl no puede preservar imágenes/macros al guardar")

# Footer
st.markdown("---")
st.markdown("""
<div style="text-align: center; color: #666; font-size: 12px;">
    🌳 Asistente de Registro de Árboles | Google Sheets o Excel
</div>
""", unsafe_allow_html=True)
